#!/usr/bin/env python3
"""Export a verified paper list to a reference library: RIS + BibTeX + by-category Excel.

Why RIS is the default: a single `.ris` file imports natively into BOTH Zotero and EndNote,
so we avoid tool-specific APIs and credential setup. `.bib` is added for LaTeX users. The
Excel is for the human to read and prune by category before importing.

Run this only on papers that passed DOI verification + the adversarial review gate. It prefers
`verified_doi` (set by verify_doi.py) over the raw `doi`, so a corrected DOI flows through.

Stdlib for RIS/BibTeX (plain text). Excel uses openpyxl if available; otherwise a .csv is written
with the same content so the step never hard-fails on a missing dependency.

Input  : verified papers (.json / .tsv / .csv). Fields: title, authors, year, journal,
         doi or verified_doi, category, abstract, doi_status.
Output : <outdir>/library.ris, library.bib, and library.xlsx (or library.csv fallback).

Usage: python export_refs.py <verified.json|tsv|csv> <outdir>
"""
import sys, os, json, csv, re, html, math, argparse

STATUS_FILL = {"verified": "C6EFCE", "candidate": "FFEB9C", "mismatch": "FFC7CE",
               "dead": "FFC7CE", "no_doi": "F2F2F2", "unverified": "F2F2F2"}   # green / yellow / red / red / grey


def clean_text(s):
    """Strip XML/JATS tags, unescape entities, collapse whitespace to one line."""
    if not s:
        return ""
    if isinstance(s, list):
        s = " ".join(str(x) for x in s)
    s = re.sub(r"<[^>]+>", "", str(s))
    s = html.unescape(s)
    return " ".join(s.split())


def loads_lenient(text):
    """Tolerant JSON parse — a weak model often wraps output in ```json fences or leaves a
    trailing comma. Strip those rather than crash the export with a raw JSONDecodeError."""
    t = text.strip()
    t = re.sub(r"^```[A-Za-z0-9]*\s*", "", t)
    t = re.sub(r"\s*```$", "", t).strip()
    try:
        return json.loads(t)
    except json.JSONDecodeError:
        return json.loads(re.sub(r",(\s*[}\]])", r"\1", t))


def read_rows(path):
    if path.lower().endswith(".json"):
        d = loads_lenient(open(path, encoding="utf-8").read())
        return d if isinstance(d, list) else d.get("papers", [])
    delim = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f, delimiter=delim)]


def author_list(row):
    """Return a list of authors from a string ('A; B' / 'A and B' / 'A, B') or an actual list."""
    a = row.get("authors", "")
    if isinstance(a, list):
        return [x for x in a if x]
    a = (a or "").strip()
    if not a:
        return []
    if ";" in a:
        return [x.strip() for x in a.split(";") if x.strip()]
    if " and " in a:
        return [x.strip() for x in a.split(" and ") if x.strip()]
    return [a]                       # keep "Last, First" style intact when no clear separator


def best_doi(row):
    return (row.get("verified_doi") or row.get("doi") or "").strip()


def trusted_doi(row):
    """Only a DOI that verify_doi CONFIRMED (doi_status == 'verified') is safe to emit as a
    trusted DOI into RIS/BibTeX/Zotero. A title-search PROPOSAL (candidate/mismatch/unverified)
    or a never-verified raw DOI must NOT land in a reference manager looking confirmed — RIS/BibTeX
    carry no status field, so an unconfirmed DOI there reads as fully trusted. Prime directive."""
    return best_doi(row) if (row.get("doi_status") == "verified") else ""


# the model's one-line Chinese key point, under any plausible field name a weak model might invent
KP_RE = re.compile(
    r"(?:key[_\s-]?point|keypoint)[_\s-]?(?:zh|cn|chinese)"   # keypoint_zh / key_point_chinese
    r"|(?:zh|cn|chinese)[_\s-]?(?:key[_\s-]?point|keypoint)"  # zh_keypoint
    r"|cn[_\s-]?(?:summary|keypoint|key[_\s-]?point)"         # cn_summary / cn_keypoint
    r"|key[_\s-]?point[_\s-]?zh",                             # key_point_zh
    re.I)


def keypoint(r):
    for k, v in r.items():
        if isinstance(v, str) and v.strip() and KP_RE.search(k):
            return v
    return ""


def split_name(author):
    """Return (surname, given) for an author string, handling the three common shapes:
    'Surname, Given' · 'Given Surname' · 'Surname Initials' (CrossRef/PubMed, e.g. 'Smith J', 'Wang YK').
    The last is the dominant form from these APIs and the one a naive last-token rule corrupts."""
    a = (author or "").strip()
    if "," in a:
        s, _, g = a.partition(",")
        return s.strip(), g.strip()
    toks = a.split()
    if len(toks) <= 1:
        return a, ""
    if re.fullmatch(r"(?:[A-Z]\.?){1,3}", toks[-1]):     # trailing initials block → first token is the surname
        return toks[0], " ".join(toks[1:])
    return toks[-1], " ".join(toks[:-1])                 # otherwise assume 'Given ... Surname'


def surname(author):
    return split_name(author)[0] or "anon"


def bib_key(row, used):
    au = author_list(row)
    k = re.sub(r"[^A-Za-z0-9]", "", surname(au[0]) if au else "anon") or "ref"
    k = f"{k}{row.get('year','')}"
    n, base = 0, k
    while k in used:
        n += 1
        suf, x = "", n
        while x:
            x, r = divmod(x - 1, 26)
            suf = chr(97 + r) + suf
        k = f"{base}{suf}"
    used.add(k)
    return k


def write_ris(rows, path):
    """RIS: imports into Zotero AND EndNote."""
    out = []
    for r in rows:
        out.append("TY  - JOUR")
        for au in author_list(r):
            if "," not in au:
                s, g = split_name(au)                # 'Smith J' → 'Smith, J' (not the corrupt 'J, Smith')
                au = f"{s}, {g}" if g else s
            out.append(f"AU  - {au}")
        if r.get("title"):   out.append(f"TI  - {clean_text(r['title'])}")
        if r.get("year"):    out.append(f"PY  - {r['year']}")
        if r.get("journal"): out.append(f"JO  - {clean_text(r['journal'])}")
        td = trusted_doi(r)
        if td:               out.append(f"DO  - {td}")                          # only a CONFIRMED DOI
        elif best_doi(r):    out.append(f"N1  - DOI待人工核对({r.get('doi_status','未核验')}): {best_doi(r)}")
        if r.get("abstract"):out.append(f"AB  - {clean_text(r['abstract'])}")
        if r.get("category"):out.append(f"KW  - {r['category']}")
        kp = keypoint(r)
        if kp:               out.append(f"N1  - 中文重点: {clean_text(kp)}")   # note → Zotero/EndNote
        out.append("ER  - ")
        out.append("")
    open(path, "w", encoding="utf-8").write("\n".join(out))


def bibtex_escape(s):
    return (s or "").replace("&", r"\&").replace("%", r"\%").replace("_", r"\_")


def write_bib(rows, path):
    used, out = set(), []
    for r in rows:
        key = bib_key(r, used)
        au = " and ".join(author_list(r))
        out.append(f"@article{{{key},")
        out.append(f"  title = {{{bibtex_escape(clean_text(r.get('title','')))}}},")
        if au:                out.append(f"  author = {{{bibtex_escape(au)}}},")
        if r.get("journal"):  out.append(f"  journal = {{{bibtex_escape(clean_text(r['journal']))}}},")
        if r.get("year"):     out.append(f"  year = {{{r['year']}}},")
        td = trusted_doi(r)
        if td:                out.append(f"  doi = {{{td}}},")                  # only a CONFIRMED DOI
        elif best_doi(r):     out.append(f"  note = {{DOI待人工核对({r.get('doi_status','未核验')}): {bibtex_escape(best_doi(r))}}},")
        out.append("}\n")
    open(path, "w", encoding="utf-8").write("\n".join(out))


# (column label, record key, width, kind). kind drives formatting; key reads the record.
# 中文重点 (key_point_zh) is the per-paper one-line Chinese key point the model writes during curate.
REPORT_COLS = [
    ("分类",      "category",             16, "plain"),
    ("标题",      "title",                52, "wrap"),
    ("作者",      "authors",              26, "plain"),
    ("年份",      "year",                  7, "center"),
    ("期刊",      "journal",              24, "plain"),
    ("DOI",       "verified_doi",         26, "doi"),
    ("核验",      "doi_status",           10, "status"),
    ("中文重点",   "key_point_zh",         46, "wrap"),
    ("创新评分",   "innovation_score",      9, "score"),
    ("创新评语",   "innovation_rationale", 36, "wrap"),
    ("影响因子",   "impact_factor",        10, "ifactor"),
    ("引用数",     "citation_count",        8, "center"),
    ("来源",      "source",               15, "plain"),
    ("摘要",      "abstract",             58, "wrap"),
]
TITLE_COL = 2   # 标题  (used for append-dedup + summary scan)
DOI_COL = 6     # DOI
STATUS_COL = 7  # 核验


def cell(r, c):
    if c == "verified_doi":
        return best_doi(r)
    if c == "key_point_zh":
        return keypoint(r)                        # fuzzy field-name match (key_point_zh / cn_summary / …)
    if c == "source":
        v = r.get("sources") or r.get("source") or ""
        return ", ".join(v) if isinstance(v, list) else (v or "")
    if c == "authors":
        a = r.get("authors", "")
        return "; ".join(a) if isinstance(a, list) else (a or "")
    v = r.get(c, "")
    return "; ".join(v) if isinstance(v, list) else ("" if v is None else v)


def doi_cell(r):
    """The DOI as it should DISPLAY/EXPORT: a CONFIRMED DOI bare; an unconfirmed/proposed one
    prefixed '待核对' so the value itself carries the warning even in a plain CSV (no color/status)."""
    d = best_doi(r)
    if not d:
        return ""
    return d if r.get("doi_status") == "verified" else f"待核对 {d}"


def _num(v):
    try:
        n = float(v)
        return n if math.isfinite(n) else None      # reject 'inf'/'nan'/'1e400' from malformed model output
    except (TypeError, ValueError):
        return None


def write_excel(rows, path):
    """Rich, color-coded, cumulative literature report (openpyxl); CSV fallback, never auto-installs.

    Columns include the per-paper 中文重点 (Chinese key point) the model writes during curate, plus
    optional 创新评分/创新评语/影响因子/引用数. Re-running appends only papers whose DOI/title aren't
    already in the file, so a library grows across several searches instead of being overwritten.
    """
    rows = sorted(rows, key=lambda r: (str(r.get("category", "zzz")), str(r.get("year", ""))))
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        csvp = os.path.splitext(path)[0] + ".csv"
        with open(csvp, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f); w.writerow([lbl for lbl, _, _, _ in REPORT_COLS])
            for r in rows:
                w.writerow([(doi_cell(r) if k == "verified_doi" else cell(r, k)) for _, k, _, _ in REPORT_COLS])
        print(f"[export] openpyxl 缺失 → 改写 CSV: {csvp}（装 openpyxl 可得着色 Excel；脚本绝不自动安装）")
        print("[export] ⚠ CSV 回退为单次写出（不累积/不去重）——装 openpyxl 才有跨搜索累积的 library.xlsx")
        return

    HEAD = PatternFill("solid", fgColor="2E74B7")
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")

    def style_row(ws, i, r):
        for j, (lbl, k, w, kind) in enumerate(REPORT_COLS, 1):
            val = cell(r, k)
            if k == "title" and not str(val).strip():
                val = "(无标题)"                       # never write a blank-title row — the append/summary scans rely on it
            if kind == "status" and not str(val).strip():
                val = "未核验"                          # missing status = never verified, not "no problem"
            cl = ws.cell(i, j, val)
            cl.alignment = Alignment(vertical="top", wrap_text=(kind == "wrap"),
                                     horizontal=("center" if kind in ("center", "score", "ifactor", "status") else "left"))
            if kind == "doi" and val:
                if r.get("doi_status") == "verified":   # only a CONFIRMED DOI gets a trusted clickable link
                    try:
                        cl.hyperlink = f"https://doi.org/{val}"
                        cl.font = Font(color="0563C1", underline="single")
                    except Exception:
                        pass
                else:
                    cl.value = f"待核对 {val}"          # the value itself warns: an unconfirmed/proposed DOI
            elif kind == "status":
                cl.fill = PatternFill("solid", fgColor=STATUS_FILL.get(val, "FFC7CE"))
            elif kind == "score":
                n = _num(val)
                if n is not None:
                    cl.fill = GREEN if n >= 8 else YELLOW if n >= 5 else RED
            elif kind == "ifactor":
                n = _num(val)
                if n is not None and n >= 10:
                    cl.fill = GREEN
                elif n is not None and n >= 5:
                    cl.fill = YELLOW

    # Append into an existing report (same header) so libraries accumulate; otherwise build fresh.
    wb = ws = None
    existing = set()
    start = 2
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            ws = wb["Literature"] if "Literature" in wb.sheetnames else wb.active
            # validate the FULL header (not just two cells): an old/foreign file whose first columns
            # happen to be 分类/标题 would otherwise pass and get appended in the new 14-col layout,
            # corrupting alignment. Require exact column count + every label.
            header_ok = (ws.max_column == len(REPORT_COLS) and
                         all(str(ws.cell(1, j).value) == lbl for j, (lbl, *_) in enumerate(REPORT_COLS, 1)))
            if header_ok:
                # scan ALL existing rows (don't stop at a blank-title row — that would land `start`
                # inside real data and overwrite it). Strip the '待核对' display prefix so a re-run
                # still dedups an unconfirmed DOI. Title-dedup only for rows that have no DOI.
                for row in range(2, ws.max_row + 1):
                    d0 = str(ws.cell(row, DOI_COL).value or "").replace("待核对", "").strip().lower()
                    t0 = str(ws.cell(row, TITLE_COL).value or "").strip().lower()
                    if not (d0 or t0):
                        continue
                    if d0:
                        existing.add(d0)
                    else:
                        existing.add(t0)
                start = ws.max_row + 1
            else:
                wb = None                          # old/foreign layout → rebuild cleanly
        except Exception:
            wb = None
    if wb is None:
        wb = Workbook(); ws = wb.active; ws.title = "Literature"
        for j, (lbl, k, w, kind) in enumerate(REPORT_COLS, 1):
            cl = ws.cell(1, j, lbl); cl.font = Font(bold=True, color="FFFFFF"); cl.fill = HEAD
            cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cl.column_letter].width = w
        ws.freeze_panes = "A2"; start = 2; existing = set()

    added = 0
    for r in rows:
        # index by BOTH the corrected (verified_doi) and raw doi so a paper whose DOI was fixed in one
        # run isn't re-added in a later run that still carries the raw DOI.
        keys = {k for k in (best_doi(r).strip().lower(), str(r.get("doi", "")).strip().lower()) if k}
        t = str(r.get("title", "")).strip().lower()
        # dedup on DOI when present; only fall back to title-dedup for DOI-less rows, so two
        # distinct same-title items (Editorial / Erratum) with different DOIs both survive.
        if (keys & existing) or (not keys and t and t in existing):
            continue
        style_row(ws, start + added, r); added += 1
        if keys:
            existing |= keys
        elif t:
            existing.add(t)
    last = start + added - 1
    if last >= 2:
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(REPORT_COLS)).column_letter}{last}"

    # Summary sheet (recreated each run) — totals by category and by DOI-verification status.
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    sm = wb.create_sheet("Summary")
    sm.column_dimensions["A"].width = 30; sm.column_dimensions["B"].width = 10
    sm.cell(1, 1, "research-to-paper 文献库").font = Font(bold=True, size=13, color="2E74B7")
    cats, stats, total = {}, {}, 0
    for row in range(2, ws.max_row + 1):
        if not (ws.cell(row, TITLE_COL).value or ws.cell(row, DOI_COL).value):
            continue
        total += 1
        c = str(ws.cell(row, 1).value or "(未分类)"); cats[c] = cats.get(c, 0) + 1
        s = str(ws.cell(row, STATUS_COL).value or "未核验"); stats[s] = stats.get(s, 0) + 1
    sm.cell(3, 1, f"共 {total} 篇  (本次 +{added})").font = Font(bold=True)
    sm.cell(5, 1, "分类").font = Font(bold=True); sm.cell(5, 2, "篇数").font = Font(bold=True)
    rr = 6
    for c, n in sorted(cats.items()):
        sm.cell(rr, 1, c); sm.cell(rr, 2, n); rr += 1
    rr += 1
    sm.cell(rr, 1, "DOI 核验").font = Font(bold=True); sm.cell(rr, 2, "篇数").font = Font(bold=True); rr += 1
    for s, n in sorted(stats.items()):
        sm.cell(rr, 1, s or "(空)"); sm.cell(rr, 2, n)
        if s in STATUS_FILL:
            sm.cell(rr, 1).fill = PatternFill("solid", fgColor=STATUS_FILL[s])
        rr += 1
    wb.save(path)
    print(f"[export] Excel: 本次 +{added} 篇，库内共 {total} 篇 → {path}")
    # diagnostic: a weak model may skip key_point_zh or write it in English — surface that, don't hide it
    n_any = sum(1 for r in rows if keypoint(r))
    n_zh = sum(1 for r in rows if any("一" <= c <= "鿿" for c in keypoint(r)))
    if rows and n_zh < len(rows):
        print(f"[export] 中文重点: {n_zh}/{len(rows)} 篇为中文（{len(rows) - n_any} 缺失, {n_any - n_zh} 疑似非中文）→ 建议补全")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("infile"); ap.add_argument("outdir")
    a = ap.parse_args()
    rows = read_rows(a.infile)
    if not rows:
        sys.exit("[export] 输入为空 / empty input")
    os.makedirs(a.outdir, exist_ok=True)
    write_ris(rows, os.path.join(a.outdir, "library.ris"))
    write_bib(rows, os.path.join(a.outdir, "library.bib"))
    write_excel(rows, os.path.join(a.outdir, "library.xlsx"))
    cats = {}
    for r in rows:
        c = r.get("category") or "(未分类)"
        if isinstance(c, list):                       # a weak model may emit category as a list — unhashable as a key
            c = ", ".join(map(str, c)) or "(未分类)"
        cats[c] = cats.get(c, 0) + 1
    print(f"[export] {len(rows)} 篇 → {a.outdir}/  (library.ris / library.bib / library.xlsx)")
    print(f"[export] 分类: {cats}")
    print("[export] 把 library.ris 导入 Zotero 或 EndNote 即可（两者都原生支持 RIS）")


if __name__ == "__main__":
    main()
